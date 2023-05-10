#!/usr/bin/env python
# coding: utf-8

# In[2]:


import requests
from bs4 import BeautifulSoup as bs
import csv
import pandas as pd
import re


# ### Scrap data kategori buku

# In[3]:


r = requests.get('http://books.toscrape.com/')

# Parsing the HTML
soup = bs(r.text, 'html.parser') 
s = soup.find('div', id_='side_categories')
content = s.find_all('a')

categories_list=[]
count = 1
for category in content:
    a = re.sub('[\t\n]', ' ', category.text).strip()
    if a == 'Books':
        pass
    else :
        d = {}
        d['Category Number'] = f'Category {count}'
        d['Category Name'] = re.sub('[\t\n]', ' ', category.text).strip()
        d['Category Website'] = ('https://books.toscrape.com/' + category.get('href'))
        d['Category Link'] = ('=HYPERLINK("https://books.toscrape.com/' + category.get('href') + '")')
        categories_list.append(d)
        count = count + 1

        
df1 = pd.DataFrame(categories_list)


# In[ ]:





# ### Scrap data semua buku secara umum

# In[4]:


URL = 'https://books.toscrape.com/catalogue/page-'

count = 1
books_list=[]
page = 1

for i in range(1,51):
    req = requests.get(URL + str(i) + '.html')
    soup = bs(req.text, 'html.parser')
    soup2 = bs(req.content, 'html.parser') 
    s = soup.find('h3')
    content = soup.find_all('h3')

    prices = soup.find_all('p', class_='price_color')
    stocks = soup.find_all('p', class_='instock availability')
    images = soup2.find_all('img')
    
    for (book,i,price,stock) in zip(content,images,prices,stocks):
        a = re.sub('[\t\n]', '', book.text).strip()
        d = {}
        d['Book Number'] = f'Book {count}'
        d['Book Name'] = re.sub('[\t\n]', '', i.get('alt')).strip()
        d['Book Cover Alt'] = i.get('alt')
        d['Book Cover'] = ('=HYPERLINK("https://books.toscrape.com/' + i.get('src').replace('..','') + '")')
        d['Book Price'] = price.text.replace('Â','')
        d['Book Availibilty'] = re.sub('[\t\n]', '', stock.text).strip()
        d['Book Link'] = ('=HYPERLINK("https://books.toscrape.com/catalogue/' + book.find('a').get('href') + '")')
        books_list.append(d)
        count = count + 1
            
df2 = pd.DataFrame(books_list)


# ### Scrap data buku - general berdasarkan kategori

# In[5]:


with pd.ExcelWriter(r'C:\Users\Mel\books1.xlsx') as writer:
    df1.to_excel(writer, sheet_name="Categories", index=False)
    df2.to_excel(writer, sheet_name="All Books", index=False)


# In[19]:


from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\Mel\books1.xlsx')


# In[20]:


ws = wb.worksheets[0]


# In[21]:


x = 2
books_listB=[]
books_lists=[]
start = 1
cek = 0
list_url=[]

def scraping(r):
    soup = bs(r.text, 'html.parser')
    soup2 = bs(r.content, 'html.parser') 
    content = soup.find_all('h3')

    pages = soup.find_all('li',class_='next')
    #         for i in pages:
    #             print(url + '/' + i.find('a').get('href'))
    prices = soup.find_all('p', class_='price_color')
    stocks = soup.find_all('p', class_='instock availability')
    images = soup2.find_all('img')

    for (book,i,price,stock) in zip(content,images,prices,stocks):
        d = {}
        d['Book Name'] = re.sub('[\t\n]', '', i.get('alt')).strip()
        d['Book Category'] = ws[('B' + str(x))].value
        d['Book Cover'] = ('=HYPERLINK("https://books.toscrape.com/' + i.get('src').replace('..','') + '")')
        d['Book Price'] = price.text.replace('Â','')
        d['Book Availibilty'] = re.sub('[\t\n]', '', stock.text).strip()
        d['Book Link'] = ('=HYPERLINK("https://books.toscrape.com/catalogue/' + book.find('a').get('href') + '")')
        books_listB.append(d)
        
    return(books_listB)
           

while ws[('C' + str(x))].value != None:
    url = ws[('C' + str(x))].value.replace('index.html','') + 'page-' + str(start) + '.html'
    r = requests.get(url)
    if r.status_code == 404:
        if cek == 0:
            url = ws[('C' + str(x))].value
            r = requests.get(url)
        print(url)
        scraping(r)

        
    if r.status_code == 200:
        while r.status_code == 200:
            url = ws[('C' + str(x))].value.replace('index.html','') + 'page-' + str(start) + '.html'
            r = requests.get(url)
            if r.status_code == 404:
                start = 1
                break
            else:
                start += 1 
            print(url)     
            scraping(r)
    x = x + 1

    


# print(books_listB)
df3 = pd.DataFrame(books_listB)


        


# In[22]:


with pd.ExcelWriter(r'C:\Users\Mel\books1.xlsx') as writer:
    df1.to_excel(writer, sheet_name="Categories", index=False)
    df2.to_excel(writer, sheet_name="All Books", index=False)
    df3.to_excel(writer, sheet_name="Books by Categories", index=False)


# In[ ]:





# ### Mengambil data informasi buku secara detail

# In[52]:


URL = 'https://books.toscrape.com/catalogue/page-'
books_detail=[]
page = 1
URL2 = ""
count = 1
for i in range(1,51):
    req = requests.get(URL + str(i) + '.html')
    soup = bs(req.text, 'html.parser')
    soup2 = bs(req.content, 'html.parser') 
    s = soup.find('h3')
    content = soup.find_all('h3')
    prices = soup.find_all('p', class_='price_color')
    stocks = soup.find_all('p', class_='instock availability')
    images = soup2.find_all('img')
#     for i in images:
#         print(i.get('src'))
    
    for (book,i,price,stock) in zip(content,images,prices,stocks):
        temp = ""
        URL2 = 'https://books.toscrape.com/catalogue/' + book.find('a').get('href')
        if URL2 in books_detail:
            pass
        else:
            req2 = requests.get(URL2)
            soup3 = bs(req2.text, 'html.parser')
            inf = soup3.find_all('th')
            dt = soup3.find_all('td')
            
            for (k,j) in zip(inf,dt):
                temp = temp + k.text + ' ' + j.text + '\n'

            
            print('\n\n', URL2)
            print(i.get('src'))
            if (soup3.find('p',class_=False)) != None:
                print(soup3.find('p',class_=False).text)
            else:
                print('Noneee')
            d = {}
            d['Book Number'] = f'Book {count}'
            d['Book Name'] = re.sub('[\t\n]', '', str(i.get('alt'))).strip()
            d['Book Cover'] = ('=HYPERLINK("https://books.toscrape.com/' + i.get('src').replace('..','') + '")')
            d['Book Price'] = price.text.replace('Â','')
            d['Book Availibilty'] = re.sub('[\t\n]', '', stock.text).strip()
            if (soup3.find('p',class_=False)) != None:
                d['Book Description'] = (soup3.find('p',class_=False).text)
            else:
                 d['Book Description'] = " "
            d['Book Information'] = temp
            d['Book Link'] = ('=HYPERLINK("https://books.toscrape.com/catalogue/' + book.find('a').get('href') + '")')
            books_detail.append(d)
            count+=1

# print(books_detail)
df4 = pd.DataFrame(books_detail)


# In[53]:


#menyimpan dataframe menjadi 1 file csv
df4.to_csv(r'C:\Users\Mel\BookDetail.csv',index=False,header=True)


# In[ ]:





# ## Menggabungkan 2 file menjadi 1

# In[4]:


from pandas import ExcelWriter
from pandas import ExcelFile


# In[5]:


dfone = pd.read_excel(r'C:\Users\Mel\books1.xlsx', sheet_name='Books by Categories')
dftwo = pd.DataFrame(pd.read_csv('BookDetail.csv'))
dftwo.drop(['Book Number'], axis=1)

dfone.head(5)


# In[6]:


dftwo.head(6)


# In[7]:


outer_join = pd.merge(dftwo, 
                      dfone, 
                      on ='Book Name', 
                      how ='outer')
outer_join.head(5)


# In[9]:


outer_join.to_csv(r'C:\Users\Mel\CombinedFile.csv',index=False,header=True)


# In[ ]:





# In[ ]:





# In[31]:


# Data Structures
import numpy  as np
import pandas as pd

# Corpus Processing
import re
import nltk.corpus
from unidecode                        import unidecode
from nltk.tokenize                    import word_tokenize
from nltk                             import SnowballStemmer
from sklearn.feature_extraction.text  import TfidfVectorizer
from sklearn.preprocessing            import normalize

# K-Means
from sklearn import cluster

# Visualization and Analysis
import matplotlib.pyplot  as plt
import matplotlib.cm      as cm
import seaborn            as sns
from sklearn.metrics                  import silhouette_samples, silhouette_score
from wordcloud                        import WordCloud


# In[32]:


data = pd.read_csv('CombinedFile.csv', encoding='utf-8')


# In[41]:


dfcluster = data[['Book Category','Book Name','Book Description']]
dfcluster.head(10)


# In[55]:


dfcluster = dfcluster.fillna(' ')


# In[ ]:





# ### Pengolahan data tanpa menggunakan stopwords removal, lemmatizing stemming dsb

# In[74]:


corpus = dfcluster['Book Description'].tolist()
corpus[9][0:500]


# In[75]:


vectorizer = TfidfVectorizer()
X = vectorizer.fit_transform(corpus)
tf_idf = pd.DataFrame(data = X.toarray(), columns=vectorizer.get_feature_names_out())

final_df = tf_idf

print("{} rows".format(final_df.shape[0]))
final_df.T.nlargest(6, 0)


# In[76]:


def run_KMeans(max_k, data):
    max_k += 1
    kmeans_results = dict()
    for k in range(2 , max_k):
        kmeans = cluster.KMeans(n_clusters = k
                               , init = 'k-means++'
                               , n_init = 10
                               , tol = 0.0001
                               , random_state = 1
                               , algorithm = 'lloyd')

        kmeans_results.update( {k : kmeans.fit(data)} )
        
    return kmeans_results


# In[77]:


def printAvg(avg_dict):
    for avg in sorted(avg_dict.keys(), reverse=True):
        print("Avg: {}\tK:{}".format(avg.round(4), avg_dict[avg]))


# In[2]:


k = 8
kmeans_results = run_KMeans(k, final_df)
print(kmeans_results)


# In[79]:


def get_top_features_cluster(tf_idf_array, prediction, n_feats):
    labels = np.unique(prediction)
    dfs = []
    for label in labels:
        id_temp = np.where(prediction==label) # indices for each cluster
        x_means = np.mean(tf_idf_array[id_temp], axis = 0) # returns average score across cluster
        sorted_means = np.argsort(x_means)[::-1][:n_feats] # indices with top 20 scores
        features = vectorizer.get_feature_names_out()
        best_features = [(features[i], x_means[i]) for i in sorted_means]
        df = pd.DataFrame(best_features, columns = ['features', 'score'])
        dfs.append(df)
    return dfs

def plotWords(dfs, n_feats):
    plt.figure(figsize=(8, 4))
    for i in range(0, len(dfs)):
        plt.title(("Most Common Words in Cluster {}".format(i)), fontsize=10, fontweight='bold')
        sns.barplot(x = 'score' , y = 'features', orient = 'h' , data = dfs[i][:n_feats])
        plt.show()


# In[1]:


best_result = 6
kmeans = kmeans_results.get(best_result)

final_df_array = final_df.to_numpy()
print(final_df_array)
# prediction = kmeans.predict(final_df)
# n_feats = 20
# dfs = get_top_features_cluster(final_df_array, prediction, n_feats)
# plotWords(dfs, 13)


# In[ ]:





# ### Pengolahan data dengan menggunakan stopwords removal, stemming lemmatizing dsb

# In[81]:


# removes a list of words (ie. stopwords) from a tokenized list.
def removeWords(listOfTokens, listOfWords):
    return [token for token in listOfTokens if token not in listOfWords]

# applies stemming to a list of tokenized words
def applyStemming(listOfTokens, stemmer):
    return [stemmer.stem(token) for token in listOfTokens]

# removes any words composed of less than 2 or more than 21 letters
def twoLetters(listOfTokens):
    twoLetterWord = []
    for token in listOfTokens:
        if len(token) <= 2 or len(token) >= 21:
            twoLetterWord.append(token)
    return twoLetterWord


# In[82]:


def processCorpus(corpus, language):   
    stopwords = nltk.corpus.stopwords.words(language)
    param_stemmer = SnowballStemmer(language)
    book_name_list = [line.rstrip('\n') for line in open('bookName.txt')] # Load .txt file line by line
    other_words = [line.rstrip('\n') for line in open('stopwords_scrapmaker.txt')] # Load .txt file line by line
    
    for document in corpus:
        index = corpus.index(document)
        corpus[index] = corpus[index].replace(u'\ufffd', '8')   # Replaces the ASCII '�' symbol with '8'
        corpus[index] = corpus[index].replace(',', '')          # Removes commas
        corpus[index] = corpus[index].rstrip('\n')              # Removes line breaks
        corpus[index] = corpus[index].casefold()                # Makes all letters lowercase
        
        corpus[index] = re.sub('\W_',' ', corpus[index])        # removes specials characters and leaves only words
        corpus[index] = re.sub("\S*\d\S*"," ", corpus[index])   # removes numbers and words concatenated with numbers IE h4ck3r. Removes road names such as BR-381.
        corpus[index] = re.sub("\S*@\S*\s?"," ", corpus[index]) # removes emails and mentions (words with @)
        corpus[index] = re.sub(r'http\S+', '', corpus[index])   # removes URLs with http
        corpus[index] = re.sub(r'www\S+', '', corpus[index])    # removes URLs with www

        listOfTokens = word_tokenize(corpus[index])
        twoLetterWord = twoLetters(listOfTokens)

        listOfTokens = removeWords(listOfTokens, stopwords)
        listOfTokens = removeWords(listOfTokens, twoLetterWord)
        listOfTokens = removeWords(listOfTokens, book_name_list)
        listOfTokens = removeWords(listOfTokens, other_words)
        
        listOfTokens = applyStemming(listOfTokens, param_stemmer)
        listOfTokens = removeWords(listOfTokens, other_words)

        corpus[index]   = " ".join(listOfTokens)
        corpus[index] = unidecode(corpus[index])

    return corpus


# In[83]:


language = 'english'
corpus = processCorpus(corpus, language)
corpus[9][0:500]


# In[84]:


vectorizer = TfidfVectorizer()
X = vectorizer.fit_transform(corpus)
tf_idf = pd.DataFrame(data = X.toarray(), columns=vectorizer.get_feature_names_out())

final_df = tf_idf

print("{} rows".format(final_df.shape[0]))
# first 6 words with highest weight on document 0:
final_df.T.nlargest(6, 0)


# In[85]:


# Running Kmeans
k = 8
kmeans_results = run_KMeans(k, final_df)
print(kmeans_results)


# In[86]:


best_result = 6
kmeans = kmeans_results.get(best_result)

final_df_array = final_df.to_numpy()
prediction = kmeans.predict(final_df)
n_feats = 20
dfs = get_top_features_cluster(final_df_array, prediction, n_feats)
plotWords(dfs, 13)


# In[87]:


# Transforms a centroids dataframe into a dictionary to be used on a WordCloud.
def centroidsDict(centroids, index):
    a = centroids.T[index].sort_values(ascending = False).reset_index().values
    centroid_dict = dict()

    for i in range(0, len(a)):
        centroid_dict.update( {a[i,0] : a[i,1]} )

    return centroid_dict

def generateWordClouds(centroids):
    wordcloud = WordCloud(max_font_size=100, background_color = 'white')
    for i in range(0, len(centroids)):
        centroid_dict = centroidsDict(centroids, i)        
        wordcloud.generate_from_frequencies(centroid_dict)

        plt.figure()
        plt.title('Cluster {}'.format(i))
        plt.imshow(wordcloud)
        plt.axis("off")
        plt.show()


# In[88]:


centroids = pd.DataFrame(kmeans.cluster_centers_)
centroids.columns = final_df.columns
generateWordClouds(centroids)


# In[ ]:




